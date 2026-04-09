"""XLSX export for results and evaluation reports.

Produces Excel workbooks with:

* **Results export** — one row per batch output with schema-driven column ordering.
* **Evaluation export** — three sheets: Summary, Confusion Matrix, Raw Data,
  plus an optional ROC Data sheet.
"""

from __future__ import annotations

import json
import logging
from pathlib import Path
from typing import Any

from openpyxl import Workbook

from llm_batch_pipeline.evaluation import EvalReport, compute_auc
from llm_batch_pipeline.logging_utils import log_event
from llm_batch_pipeline.schema_loader import load_schema_class

logger = logging.getLogger("llm_batch_pipeline.export")


# ---------------------------------------------------------------------------
# Results XLSX
# ---------------------------------------------------------------------------


def export_results_xlsx(
    rows: list[dict[str, Any]],
    output_path: Path,
    *,
    schema_path: Path | None = None,
) -> None:
    """Export validated result rows to an XLSX workbook."""
    columns = _resolve_columns(rows, schema_path)

    wb = Workbook()
    ws = wb.active
    ws.title = "Results"

    # Header
    for col_idx, col_name in enumerate(columns, 1):
        ws.cell(row=1, column=col_idx, value=col_name)

    # Data
    for row_idx, row in enumerate(rows, 2):
        for col_idx, col_name in enumerate(columns, 1):
            value = row.get(col_name)
            ws.cell(row=row_idx, column=col_idx, value=_coerce_cell(value))

    # Formatting
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    log_event(logger, f"Exported {len(rows)} rows to {output_path}", step="export", status="ok", rows=len(rows))


# ---------------------------------------------------------------------------
# Evaluation XLSX
# ---------------------------------------------------------------------------


def export_evaluation_xlsx(
    report: EvalReport,
    output_path: Path,
    *,
    schema_path: Path | None = None,
) -> None:
    """Export evaluation report to a multi-sheet XLSX workbook."""
    wb = Workbook()

    # Sheet 1: Summary
    _write_summary_sheet(wb.active, report)

    # Sheet 2: Confusion Matrix
    _write_confusion_sheet(wb.create_sheet("Confusion Matrix"), report)

    # Sheet 3: Raw Data
    _write_raw_data_sheet(wb.create_sheet("Raw Data"), report, schema_path)

    # Sheet 4: ROC Data (if available)
    if report.roc_data:
        _write_roc_sheet(wb.create_sheet("ROC Data"), report)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    log_event(logger, f"Exported evaluation to {output_path}", step="export", status="ok")


# ---------------------------------------------------------------------------
# Sheet writers
# ---------------------------------------------------------------------------


def _write_summary_sheet(ws: Any, report: EvalReport) -> None:
    ws.title = "Summary"

    # Overall metrics
    metrics = [
        ("Accuracy", report.accuracy),
        ("Macro Precision", report.macro_precision),
        ("Macro Recall", report.macro_recall),
        ("Macro F1", report.macro_f1),
        ("Weighted Precision", report.weighted_precision),
        ("Weighted Recall", report.weighted_recall),
        ("Weighted F1", report.weighted_f1),
        ("Total Predictions", len(report.rows)),
    ]

    if report.roc_data:
        metrics.append(("AUC", compute_auc(report.roc_data)))

    for row_idx, (name, value) in enumerate(metrics, 1):
        ws.cell(row=row_idx, column=1, value=name)
        ws.cell(row=row_idx, column=2, value=round(value, 4) if isinstance(value, float) else value)

    # Per-class table
    start_row = len(metrics) + 3
    headers = ["Class", "Precision", "Recall", "F1", "Support", "TP", "FP", "FN"]
    for col_idx, h in enumerate(headers, 1):
        ws.cell(row=start_row, column=col_idx, value=h)

    for i, cm in enumerate(report.per_class):
        r = start_row + 1 + i
        vals = [cm.label, round(cm.precision, 4), round(cm.recall, 4), round(cm.f1, 4), cm.support, cm.tp, cm.fp, cm.fn]
        for col_idx, v in enumerate(vals, 1):
            ws.cell(row=r, column=col_idx, value=v)


def _write_confusion_sheet(ws: Any, report: EvalReport) -> None:
    # Header row: blank + labels
    for col_idx, label in enumerate(report.labels, 2):
        ws.cell(row=1, column=col_idx, value=f"Pred: {label}")

    # Data rows
    for row_idx, (label, row_data) in enumerate(zip(report.labels, report.confusion, strict=True), 2):
        ws.cell(row=row_idx, column=1, value=f"True: {label}")
        for col_idx, count in enumerate(row_data, 2):
            ws.cell(row=row_idx, column=col_idx, value=count)


def _write_raw_data_sheet(ws: Any, report: EvalReport, schema_path: Path | None) -> None:
    base_cols = ["filename", "ground_truth", "predicted", "confidence", "correct"]

    # Extra schema columns
    extra_cols: list[str] = []
    if schema_path and schema_path.is_file():
        try:
            cls = load_schema_class(schema_path)
            extra_cols = [
                name for name in cls.model_fields if name not in ("label", "classification", "confidence", "certainty")
            ]
        except Exception:  # noqa: BLE001
            pass

    all_cols = base_cols + extra_cols

    for col_idx, col in enumerate(all_cols, 1):
        ws.cell(row=1, column=col_idx, value=col)

    for row_idx, pred_row in enumerate(report.rows, 2):
        ws.cell(row=row_idx, column=1, value=pred_row.custom_id)
        ws.cell(row=row_idx, column=2, value=pred_row.ground_truth)
        ws.cell(row=row_idx, column=3, value=pred_row.predicted)
        ws.cell(row=row_idx, column=4, value=pred_row.confidence)
        ws.cell(row=row_idx, column=5, value=pred_row.ground_truth == pred_row.predicted)

        # Extra columns from raw output
        if pred_row.raw_output:
            for col_idx, col_name in enumerate(extra_cols, len(base_cols) + 1):
                val = pred_row.raw_output.get(col_name)
                ws.cell(row=row_idx, column=col_idx, value=_coerce_cell(val))

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def _write_roc_sheet(ws: Any, report: EvalReport) -> None:
    headers = ["FPR", "TPR", "Threshold"]
    for col_idx, h in enumerate(headers, 1):
        ws.cell(row=1, column=col_idx, value=h)

    for row_idx, point in enumerate(report.roc_data, 2):
        ws.cell(row=row_idx, column=1, value=point["fpr"])
        ws.cell(row=row_idx, column=2, value=point["tpr"])
        ws.cell(row=row_idx, column=3, value=point["threshold"])


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------


def _resolve_columns(rows: list[dict[str, Any]], schema_path: Path | None) -> list[str]:
    """Determine column order: schema-driven first, then extras."""
    ordered: list[str] = ["filename"]

    if schema_path and schema_path.is_file():
        try:
            cls = load_schema_class(schema_path)
            for name, field_info in cls.model_fields.items():
                alias = getattr(field_info, "alias", None)
                ordered.append(alias or name)
        except Exception:  # noqa: BLE001
            pass

    # Discover extra keys from data
    seen = set(ordered)
    for row in rows:
        for key in row:
            if key not in seen:
                ordered.append(key)
                seen.add(key)

    return ordered


def _coerce_cell(value: Any) -> Any:
    """Convert complex types to strings for XLSX cells."""
    if isinstance(value, (list, dict)):
        return json.dumps(value, ensure_ascii=False, default=str)
    if isinstance(value, str):
        # Strip XML 1.0 illegal characters
        return _sanitise_for_xlsx(value)
    return value


def _sanitise_for_xlsx(text: str) -> str:
    """Remove characters illegal in XML 1.0 (used by XLSX)."""
    return "".join(
        c
        for c in text
        if c == "\t" or c == "\n" or c == "\r" or ("\x20" <= c <= "\ud7ff") or ("\ue000" <= c <= "\ufffd")
    )
